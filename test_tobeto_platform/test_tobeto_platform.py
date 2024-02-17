from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c


class TestTobetoPlatform():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def get_data_survey():
    excel = openpyxl.load_workbook(c.VALID_LOGIN_XLSX)
    sheet = excel["Sayfa1"] 
    data = []
    for i in range(2,3):
        email = sheet.cell(i,1).value
        password = sheet.cell(i,2).value
        data.append((email,password))
      
    return data


  @pytest.mark.parametrize("email,password",get_data_survey())
  def test_survey_button(self,email,password): 
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
    valid_login = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert valid_login.text == c.VALID_LOGIN
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH, c.SURVEY_BUTTON_XPATH))).click()
    survey_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.SURVEY_MESSAGE_XPATH))) 
    assert survey_message.text == c.SURVEY_MESSAGE
    sleep(3)


  @pytest.mark.parametrize("email,password",get_data_survey())
  def test_catalog_button(self,email,password): 
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
    valid_login = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert valid_login.text == c.VALID_LOGIN
    WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, c.CATALOG_BUTTON_XPATH))).click()
    WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.ID,c.EDUCATION_ID))).send_keys("Tobeto")
    WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, c.SEARCH_BUTTON))).click()
    catalog_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.CATALOG_MESSAGE_XPATH))) 
    assert catalog_message.text == c.CATALOG_MESSAGE
    sleep(3)
