from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c


class TestTobetoPlatform():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def get_data_survey():
    excel = openpyxl.load_workbook(c.SURVEY_XLSX)
    sheet = excel["Sayfa1"] 
    data = []
    for i in range(2,3):
        email = sheet.cell(i,1).value
        password = sheet.cell(i,2).value
        data.append((email,password,))
      
    return data


  @pytest.mark.parametrize("email,password",get_data_survey())
  def test_invalid_information(self,email,password): 
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
    valid_login = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert valid_login.text == c.VALID_LOGIN
    sleep(3)
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.MY_PROFILE_XPATH))).click()
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.EDIT_BUTTON_XPATH))).click()
    tc_message = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.XPATH, c.TC_MESSAGE_XPATH))) 
    assert tc_message.text == c.TC__MESSAGE
    sleep(3)


  @pytest.mark.parametrize("email,password",get_data_survey())
  def test_empty(self,email,password): 
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
    valid_login = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert valid_login.text == c.VALID_LOGIN
    sleep(3)
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.MY_PROFILE_XPATH))).click()
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.EDIT_BUTTON_XPATH))).click()
    self.driver.execute_script("window.scrollTo(0,750)")
    sleep(3)
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.SAVE_BUTTON_XPATH))).click()
    sleep(3)
    error_message = WebDriverWait(self.driver,40).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH))) 
    assert error_message.text == c.ERROR_MESSAGE
    sleep(3)


  @pytest.mark.parametrize("email,password",get_data_survey())
  def test_my_experiences(self,email,password): 
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
    valid_login = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert valid_login.text == c.VALID_LOGIN
    sleep(3)
    WebDriverWait(self.driver, 40).until(ec.visibility_of_element_located((By.XPATH, c.MY_PROFILE_XPATH))).click()
    sleep(3)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH, c.EDIT_BUTTON_XPATH))).click()
    sleep(3)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH, c.EXPERIENCE_BUTTON_XPATH))).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,c.ORGANIZATION_NAME_XPATH))).send_keys("Tobeto")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,c.POSITION_XPATH))).send_keys("Yazılım Test Uzmanı")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,c.SECTOR_XPATH))).send_keys("Yazılım")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,c.SECTOR_XPATH))).click()
    sleep(3)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,c.CITY_XPATH))).send_keys("İstanbul")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.XPATH,c.STARTUP_XPATH))).send_keys("01.01.2024")
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.CONTINUED_XPATH))).click()
    WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.XPATH, c.SAVE_BUTTON_XPATH))).click()
    experiense_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.EXPERIENCE_MESSAGE_XPATH))) 
    assert experiense_message.text == c.EXPERIENCE_MESSAGE
    sleep(3)








