from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c


class TestRegister():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def get_data_register():
      excel = openpyxl.load_workbook(c.REGISTER_XLSX)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          password = sheet.cell(i,4).value
          passwordAgain = sheet.cell(i,5).value
          phoneNumber = sheet.cell(i,6).value
          data.append((firstName,lastName,email,password,passwordAgain,phoneNumber))

      return data
  
  def get_data_without_phone():
      excel = openpyxl.load_workbook(c.REGISTER_PHONE_XLSX)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          password = sheet.cell(i,4).value
          passwordAgain = sheet.cell(i,5).value
          data.append((firstName,lastName,email,password,passwordAgain))

      return data
  
  def get_data_without_password():
      excel = openpyxl.load_workbook(c.REGISTER_PASSWORD_XLSX)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          firstName = sheet.cell(i,1).value
          lastName = sheet.cell(i,2).value
          email = sheet.cell(i,3).value
          phoneNumber= sheet.cell(i,4).value
          
          data.append((firstName,lastName,email,phoneNumber))

      return data


  def test_invalidEmail(self):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "email"))).send_keys("abcd")
    errorMessage = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.INVALID_EMAIL_XPATH))) 
    assert errorMessage.text == c.INVALID_EMAIL 
  

  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain,phoneNumber", get_data_register()) 
  def test_invalidRegister(self,firstName,lastName,email,password,passwordAgain,phoneNumber):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME))).send_keys(firstName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME))).send_keys(lastName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME))).send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_AGAIN_NAME))).send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    sleep(1)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH))).click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID))).send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
    self.driver.switch_to.default_content()
    sleep(30)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH))).click()
    cant_register = WebDriverWait(self.driver, 20).until(ec.visibility_of_element_located((By.CLASS_NAME, "toast-body")))
    assert cant_register.text == c.CANT_REGISTER


  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain", get_data_without_phone()) 
  def test_minPhoneNumber(self,firstName,lastName,email,password,passwordAgain):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME))).send_keys(firstName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME))).send_keys(lastName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME))).send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_AGAIN_NAME))).send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    sleep(1)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH))).click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID))).send_keys("12345")
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
    self.driver.switch_to.default_content()
    sleep(30)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH))).click()
    min_phone_number = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.MIN_PHONE_NUMBER_XPATH)))
    assert min_phone_number.text == c.MIN_PHONE_NUMBER

  @pytest.mark.parametrize("firstName,lastName,email,password,passwordAgain", get_data_without_phone()) 
  def test_maxPhoneNumber(self,firstName,lastName,email,password,passwordAgain):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME))).send_keys(firstName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME))).send_keys(lastName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME))).send_keys(password)
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_AGAIN_NAME))).send_keys(passwordAgain)
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH))).click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID))).send_keys("12345678911")
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
    self.driver.switch_to.default_content()
    sleep(30)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH))).click()
    WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.MAX_PHONE_NUMBER_XPATH)))
    max_phone_number = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, c.MAX_PHONE_NUMBER_XPATH)))
    assert max_phone_number.text == c.MAX_PHONE_NUMBER
  

  def test_Missing_Password(self):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME))).send_keys(c.FIRSTNAME)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME))).send_keys(c.LASTNAME)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "email"))).send_keys("xcvbg@gmail.com")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,"password"))).send_keys("12345")
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "passwordAgain"))).send_keys("12345")
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH))).click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID))).send_keys(c.PHONE)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
    self.driver.switch_to.default_content()
    sleep(30)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH))).click()
    missing_password = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CLASS_NAME , "toast-body")))
    assert missing_password.text == c.MISSING_PASSWORD


  @pytest.mark.parametrize("firstName,lastName,email,phoneNumber", get_data_without_password()) 
  def test_Unmatched_Password(self,firstName,lastName,email,phoneNumber):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME))).send_keys(firstName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME))).send_keys(lastName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,"password"))).send_keys("12345678")
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "passwordAgain"))).send_keys("123456")
    self.driver.execute_script("window.scrollTo(0,500)")
    sleep(5)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH))).click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID))).send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
    self.driver.switch_to.default_content()
    sleep(30)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH))).click()
    unmatched_password = self.driver.find_element(By.CLASS_NAME , "toast-body")
    assert unmatched_password.text == c.UNMATCHED_PASSWORD


  @pytest.mark.parametrize("firstName,lastName,email,phoneNumber", get_data_without_password()) 
  def test_Invalid_Informations(self,firstName,lastName,email,phoneNumber):
    self.driver.find_element(By.LINK_TEXT, c.REGISTER).click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.FIRSTNAME_NAME))).send_keys(firstName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.LASTNAME_NAME))).send_keys(lastName)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME,"password"))).send_keys("12345")
    self.driver.execute_script("window.scrollTo(0,500)")
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, "passwordAgain"))).send_keys("12345")
    self.driver.execute_script("window.scrollTo(0,500)")
    sleep(5)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REGISTER_XPATH))).click()
    self.driver.find_element(By.NAME, "contact").click()
    self.driver.find_element(By.NAME, "membershipContrat").click()
    self.driver.find_element(By.NAME, "emailConfirmation").click()
    self.driver.find_element(By.NAME, "phoneConfirmation").click()
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, c.PHONE_NUMBER_ID))).send_keys(phoneNumber)
    WebDriverWait(self.driver, 10).until(ec.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[starts-with(@name, 'a-') and starts-with(@src, 'https://www.google.com/recaptcha')]")))
    WebDriverWait(self.driver, 20).until(ec.element_to_be_clickable((By.CLASS_NAME, "recaptcha-checkbox-border"))).click()
    self.driver.switch_to.default_content()
    sleep(30)
    WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.FINAL_REGISTER_XPATH))).click()
    invalid_informations = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CLASS_NAME , "toast-body")))
    assert invalid_informations.text == c.INVALID_INFORMATIONS