from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c

class TestTobetoLogin():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.get(c.BASE_URL)
    self.driver.maximize_window()
    
  
  def teardown_method(self):
    self.driver.quit()


  def get_data_valid():
      excel = openpyxl.load_workbook(c.VALID_LOGIN_XLSX)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          email = sheet.cell(i,1).value
          password = sheet.cell(i,2).value
          data.append((email,password))
      
      return data
  
  def get_data_invalid():
      excel = openpyxl.load_workbook(c.INVALID_LOGIN_XLSX)
      sheet = excel["Sheet1"] 
      rows = sheet.max_row 
      data = []
      for i in range(2,rows+1):
          email = sheet.cell(i,1).value
          password = sheet.cell(i,2).value
          data.append((email,password))
      
      return data

  
  @pytest.mark.parametrize("email,password",[("","")])
  def test_empty_login(self,email,password):
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
    error_email_message = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMPTY_MESSAGE_EMAIL_XPATH))) 
    assert error_email_message.text == c.EMPTY_MESSAGE
    error_password_message = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMPTY_MESSAGE_PASSWORD_XPATH))) 
    assert error_password_message.text == c.EMPTY_MESSAGE, self.driver.save_screenshot("test_login/test_empty_login.png")

  @pytest.mark.parametrize("email,password",get_data_valid())
  def test_valid_Login(self,email,password):
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
    WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME))).send_keys(password)
    self.driver.find_element(By.XPATH, c.LOGIN_BUTTON_XPATH).click()
    valid_login = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, c.VALID_LOGIN_XPATH)))
    assert valid_login.text == c.VALID_LOGIN
    self.driver.save_screenshot("./valid.png")
  

  @pytest.mark.parametrize("email,password",get_data_invalid()) 
  def test_invalid_login(self,email,password):
   WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.EMAIL_NAME))).send_keys(email)
   WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.NAME, c.PASSWORD_NAME))).send_keys(password)
   self.driver.find_element(By.XPATH, c.LOGIN_BUTTON_XPATH).click()
   invalid_login = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.INVALID_LOGIN_XPATH)))
   assert invalid_login.text == c.INVALID_LOGIN
   self.driver.save_screenshot("./invalid.png")
